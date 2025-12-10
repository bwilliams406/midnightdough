"""
Midnight Dough Cookie Calculator Generator
Creates a comprehensive xlsx spreadsheet for cookie production planning

Features:
- Ingredient price database
- All cookie recipes with precise measurements
- Order calculator with quantity/size scaling
- Automatic cost calculations
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation

def create_cookie_calculator():
    wb = Workbook()
    
    # =========================================================================
    # STYLING
    # =========================================================================
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    subheader_fill = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")
    money_fill = PatternFill(start_color="e8f5e9", end_color="e8f5e9", fill_type="solid")
    input_fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # =========================================================================
    # SHEET 1: INGREDIENT PRICE DATABASE
    # =========================================================================
    ws_prices = wb.active
    ws_prices.title = "Ingredient Prices"
    
    # Header
    ws_prices.merge_cells('A1:F1')
    ws_prices['A1'] = "🍪 MIDNIGHT DOUGH - INGREDIENT PRICE DATABASE"
    ws_prices['A1'].font = Font(color="FFFFFF", bold=True, size=16)
    ws_prices['A1'].fill = header_fill
    ws_prices['A1'].alignment = Alignment(horizontal='center')
    
    ws_prices.merge_cells('A2:F2')
    ws_prices['A2'] = "Edit Package Price - Price per Unit calculates automatically"
    ws_prices['A2'].font = Font(italic=True, size=10)
    ws_prices['A2'].alignment = Alignment(horizontal='center')
    
    # Column headers
    price_headers = ["Ingredient", "Unit", "Package Size", "Package Price ($)", "Price per Unit", "Notes"]
    for col, header in enumerate(price_headers, 1):
        cell = ws_prices.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Ingredient data - NO formulas in tuple, we'll add them separately
    ingredients = [
        ("Unsalted Butter", "g", 454, 5.99, "Standard 1lb block"),
        ("Vegetable Shortening", "g", 1360, 6.49, "Crisco 48oz"),
        ("Granulated Sugar", "g", 1814, 4.29, "4lb bag"),
        ("Brown Sugar (Light)", "g", 907, 3.99, "2lb bag"),
        ("Brown Sugar (Dark)", "g", 907, 3.99, "2lb bag"),
        ("Powdered Sugar", "g", 907, 2.99, "2lb bag"),
        ("All-Purpose Flour", "g", 2268, 4.49, "5lb bag"),
        ("Cornstarch", "g", 454, 2.99, "1lb box"),
        ("Dutch-Process Cocoa", "g", 227, 8.99, "8oz premium"),
        ("Old-Fashioned Oats", "g", 1134, 4.99, "40oz container"),
        ("Baking Powder", "g", 283, 3.49, "10oz can"),
        ("Baking Soda", "g", 454, 1.99, "1lb box"),
        ("Cream of Tartar", "g", 85, 6.99, "3oz jar"),
        ("Eggs (large)", "each", 12, 4.99, "Dozen"),
        ("Egg Yolk", "each", 12, 4.99, "From dozen"),
        ("Sour Cream", "g", 454, 2.99, "16oz container"),
        ("Cream Cheese", "g", 227, 3.49, "8oz block"),
        ("Vanilla Extract", "g", 118, 8.99, "4oz pure vanilla"),
        ("Lemon Juice (fresh)", "tbsp", 3, 0.50, "Per lemon ~3tbsp"),
        ("Lemon Zest", "tbsp", 1, 0.17, "Per lemon ~1tbsp"),
        ("Lemon Extract", "g", 59, 4.99, "2oz bottle"),
        ("Espresso Powder", "g", 57, 7.99, "2oz jar"),
        ("Ground Cinnamon", "g", 68, 4.99, "2.4oz jar"),
        ("Nutmeg", "g", 62, 5.99, "2.2oz jar"),
        ("Salt (Kosher)", "g", 1361, 3.99, "3lb box"),
        ("Flaky Sea Salt", "g", 113, 6.99, "4oz Maldon"),
        ("Chocolate Chips", "g", 340, 4.99, "12oz bag"),
        ("Dark Chocolate Chunks", "g", 283, 6.99, "10oz premium"),
        ("Raisins", "g", 425, 4.49, "15oz box"),
        ("Walnuts (chopped)", "g", 227, 7.99, "8oz bag"),
        ("Pecans (chopped)", "g", 227, 9.99, "8oz bag"),
    ]
    
    # Write ingredient data with PROPER formulas
    for i, ing in enumerate(ingredients):
        row = 5 + i
        # Ingredient name
        ws_prices.cell(row=row, column=1, value=ing[0]).border = thin_border
        # Unit
        ws_prices.cell(row=row, column=2, value=ing[1]).border = thin_border
        # Package size
        cell_size = ws_prices.cell(row=row, column=3, value=ing[2])
        cell_size.border = thin_border
        # Package price (editable)
        cell_price = ws_prices.cell(row=row, column=4, value=ing[3])
        cell_price.border = thin_border
        cell_price.fill = input_fill
        cell_price.number_format = '$#,##0.00'
        # Price per unit - FORMULA
        cell_ppu = ws_prices.cell(row=row, column=5)
        cell_ppu.value = f"=D{row}/C{row}"
        cell_ppu.border = thin_border
        cell_ppu.fill = money_fill
        cell_ppu.number_format = '$#,##0.0000'
        # Notes
        ws_prices.cell(row=row, column=6, value=ing[4]).border = thin_border
    
    # Set column widths
    ws_prices.column_dimensions['A'].width = 25
    ws_prices.column_dimensions['B'].width = 10
    ws_prices.column_dimensions['C'].width = 15
    ws_prices.column_dimensions['D'].width = 18
    ws_prices.column_dimensions['E'].width = 18
    ws_prices.column_dimensions['F'].width = 25
    
    # =========================================================================
    # SHEET 2: RECIPE DATABASE
    # =========================================================================
    ws_recipes = wb.create_sheet("Recipe Database")
    
    # Header
    ws_recipes.merge_cells('A1:H1')
    ws_recipes['A1'] = "🍪 MIDNIGHT DOUGH - RECIPE DATABASE"
    ws_recipes['A1'].font = Font(color="FFFFFF", bold=True, size=16)
    ws_recipes['A1'].fill = header_fill
    ws_recipes['A1'].alignment = Alignment(horizontal='center')
    
    recipe_headers = ["Recipe Name", "Base Yield", "Base Size (g)", "Total Dough (g)", "Ingredient", "Amount", "Unit", "Category"]
    for col, header in enumerate(recipe_headers, 1):
        cell = ws_recipes.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # All recipes with ingredients
    recipes_data = [
        # Sugar Cookies
        ("Sugar Cookie", 24, 48, 1152, "Unsalted Butter", 170, "g", "Wet"),
        ("Sugar Cookie", 24, 48, 1152, "Granulated Sugar", 250, "g", "Wet"),
        ("Sugar Cookie", 24, 48, 1152, "Eggs (large)", 1, "each", "Wet"),
        ("Sugar Cookie", 24, 48, 1152, "Vanilla Extract", 8, "g", "Wet"),
        ("Sugar Cookie", 24, 48, 1152, "Sour Cream", 30, "g", "Wet"),
        ("Sugar Cookie", 24, 48, 1152, "All-Purpose Flour", 345, "g", "Dry"),
        ("Sugar Cookie", 24, 48, 1152, "Cornstarch", 8, "g", "Dry"),
        ("Sugar Cookie", 24, 48, 1152, "Baking Powder", 5, "g", "Dry"),
        ("Sugar Cookie", 24, 48, 1152, "Baking Soda", 3, "g", "Dry"),
        ("Sugar Cookie", 24, 48, 1152, "Salt (Kosher)", 4, "g", "Dry"),
        ("Sugar Cookie", 24, 48, 1152, "Granulated Sugar", 50, "g", "Rolling"),
        
        # Snickerdoodle
        ("Snickerdoodle", 24, 50, 1200, "All-Purpose Flour", 375, "g", "Dry"),
        ("Snickerdoodle", 24, 50, 1200, "Cream of Tartar", 6, "g", "Dry"),
        ("Snickerdoodle", 24, 50, 1200, "Baking Soda", 5, "g", "Dry"),
        ("Snickerdoodle", 24, 50, 1200, "Ground Cinnamon", 4, "g", "Dry"),
        ("Snickerdoodle", 24, 50, 1200, "Salt (Kosher)", 3, "g", "Dry"),
        ("Snickerdoodle", 24, 50, 1200, "Unsalted Butter", 226, "g", "Wet"),
        ("Snickerdoodle", 24, 50, 1200, "Granulated Sugar", 267, "g", "Wet"),
        ("Snickerdoodle", 24, 50, 1200, "Eggs (large)", 1, "each", "Wet"),
        ("Snickerdoodle", 24, 50, 1200, "Egg Yolk", 1, "each", "Wet"),
        ("Snickerdoodle", 24, 50, 1200, "Vanilla Extract", 8, "g", "Wet"),
        ("Snickerdoodle", 24, 50, 1200, "Granulated Sugar", 70, "g", "Topping"),
        ("Snickerdoodle", 24, 50, 1200, "Ground Cinnamon", 3, "g", "Topping"),
        
        # Dark Chocolate Chocolate Chip
        ("Dark Chocolate Chip", 10, 115, 1150, "Unsalted Butter", 226, "g", "Wet"),
        ("Dark Chocolate Chip", 10, 115, 1150, "Brown Sugar (Dark)", 220, "g", "Wet"),
        ("Dark Chocolate Chip", 10, 115, 1150, "Granulated Sugar", 100, "g", "Wet"),
        ("Dark Chocolate Chip", 10, 115, 1150, "Eggs (large)", 2, "each", "Wet"),
        ("Dark Chocolate Chip", 10, 115, 1150, "Vanilla Extract", 8, "g", "Wet"),
        ("Dark Chocolate Chip", 10, 115, 1150, "All-Purpose Flour", 280, "g", "Dry"),
        ("Dark Chocolate Chip", 10, 115, 1150, "Dutch-Process Cocoa", 60, "g", "Dry"),
        ("Dark Chocolate Chip", 10, 115, 1150, "Baking Soda", 5, "g", "Dry"),
        ("Dark Chocolate Chip", 10, 115, 1150, "Salt (Kosher)", 4, "g", "Dry"),
        ("Dark Chocolate Chip", 10, 115, 1150, "Espresso Powder", 2, "g", "Dry"),
        ("Dark Chocolate Chip", 10, 115, 1150, "Dark Chocolate Chunks", 255, "g", "Mix-in"),
        ("Dark Chocolate Chip", 10, 115, 1150, "Flaky Sea Salt", 2, "g", "Topping"),
        
        # Chocolate Chip (Double Batch)
        ("Chocolate Chip", 48, 55, 2640, "Granulated Sugar", 300, "g", "Wet"),
        ("Chocolate Chip", 48, 55, 2640, "Brown Sugar (Light)", 400, "g", "Wet"),
        ("Chocolate Chip", 48, 55, 2640, "Unsalted Butter", 454, "g", "Wet"),
        ("Chocolate Chip", 48, 55, 2640, "Vanilla Extract", 26, "g", "Wet"),
        ("Chocolate Chip", 48, 55, 2640, "Eggs (large)", 4, "each", "Wet"),
        ("Chocolate Chip", 48, 55, 2640, "All-Purpose Flour", 750, "g", "Dry"),
        ("Chocolate Chip", 48, 55, 2640, "Baking Soda", 9, "g", "Dry"),
        ("Chocolate Chip", 48, 55, 2640, "Salt (Kosher)", 9, "g", "Dry"),
        ("Chocolate Chip", 48, 55, 2640, "Chocolate Chips", 1020, "g", "Mix-in"),
        
        # Lemon Sugar (Double Batch)
        ("Lemon Sugar", 44, 30, 1320, "All-Purpose Flour", 500, "g", "Dry"),
        ("Lemon Sugar", 44, 30, 1320, "Baking Soda", 5, "g", "Dry"),
        ("Lemon Sugar", 44, 30, 1320, "Baking Powder", 5, "g", "Dry"),
        ("Lemon Sugar", 44, 30, 1320, "Salt (Kosher)", 6, "g", "Dry"),
        ("Lemon Sugar", 44, 30, 1320, "Unsalted Butter", 454, "g", "Wet"),
        ("Lemon Sugar", 44, 30, 1320, "Granulated Sugar", 400, "g", "Wet"),
        ("Lemon Sugar", 44, 30, 1320, "Eggs (large)", 2, "each", "Wet"),
        ("Lemon Sugar", 44, 30, 1320, "Lemon Juice (fresh)", 4, "tbsp", "Wet"),
        ("Lemon Sugar", 44, 30, 1320, "Lemon Zest", 2, "tbsp", "Wet"),
        ("Lemon Sugar", 44, 30, 1320, "Vanilla Extract", 8, "g", "Wet"),
        ("Lemon Sugar", 44, 30, 1320, "Granulated Sugar", 100, "g", "Rolling"),
        
        # Oatmeal Raisin
        ("Oatmeal Raisin", 24, 50, 1200, "Unsalted Butter", 226, "g", "Wet"),
        ("Oatmeal Raisin", 24, 50, 1200, "Brown Sugar (Light)", 200, "g", "Wet"),
        ("Oatmeal Raisin", 24, 50, 1200, "Granulated Sugar", 100, "g", "Wet"),
        ("Oatmeal Raisin", 24, 50, 1200, "Eggs (large)", 2, "each", "Wet"),
        ("Oatmeal Raisin", 24, 50, 1200, "Vanilla Extract", 13, "g", "Wet"),
        ("Oatmeal Raisin", 24, 50, 1200, "All-Purpose Flour", 190, "g", "Dry"),
        ("Oatmeal Raisin", 24, 50, 1200, "Baking Soda", 5, "g", "Dry"),
        ("Oatmeal Raisin", 24, 50, 1200, "Ground Cinnamon", 4, "g", "Dry"),
        ("Oatmeal Raisin", 24, 50, 1200, "Salt (Kosher)", 3, "g", "Dry"),
        ("Oatmeal Raisin", 24, 50, 1200, "Old-Fashioned Oats", 240, "g", "Mix-in"),
        ("Oatmeal Raisin", 24, 50, 1200, "Raisins", 190, "g", "Mix-in"),
        ("Oatmeal Raisin", 24, 50, 1200, "Walnuts (chopped)", 60, "g", "Mix-in"),
    ]
    
    row = 4
    recipe_fill_colors = {
        "Sugar Cookie": "fff3e0",
        "Snickerdoodle": "fce4ec",
        "Dark Chocolate Chip": "3e2723",
        "Chocolate Chip": "8d6e63",
        "Lemon Sugar": "fffde7",
        "Oatmeal Raisin": "efebe9"
    }
    
    for data in recipes_data:
        fill_color = recipe_fill_colors.get(data[0], "ffffff")
        text_color = "FFFFFF" if data[0] == "Dark Chocolate Chip" else "000000"
        
        for col, val in enumerate(data, 1):
            cell = ws_recipes.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            if text_color == "FFFFFF":
                cell.font = Font(color=text_color)
        row += 1
    
    # Set column widths
    ws_recipes.column_dimensions['A'].width = 22
    ws_recipes.column_dimensions['B'].width = 12
    ws_recipes.column_dimensions['C'].width = 14
    ws_recipes.column_dimensions['D'].width = 15
    ws_recipes.column_dimensions['E'].width = 25
    ws_recipes.column_dimensions['F'].width = 10
    ws_recipes.column_dimensions['G'].width = 8
    ws_recipes.column_dimensions['H'].width = 12
    
    # =========================================================================
    # SHEET 3: BATCH CALCULATOR (FIXED!)
    # =========================================================================
    ws_batch = wb.create_sheet("Batch Calculator")
    
    cookie_types_str = '"Sugar Cookie,Snickerdoodle,Dark Chocolate Chip,Chocolate Chip,Lemon Sugar,Oatmeal Raisin"'
    
    # Header
    ws_batch.merge_cells('A1:H1')
    ws_batch['A1'] = "📊 BATCH CALCULATOR - Scale Any Recipe"
    ws_batch['A1'].font = Font(color="FFFFFF", bold=True, size=16)
    ws_batch['A1'].fill = header_fill
    ws_batch['A1'].alignment = Alignment(horizontal='center')
    
    # Instructions
    ws_batch.merge_cells('A2:H2')
    ws_batch['A2'] = "Select cookie type, enter quantity & size. Ingredients scale automatically with costs!"
    ws_batch['A2'].font = Font(italic=True, size=10)
    ws_batch['A2'].alignment = Alignment(horizontal='center')
    
    # Input section
    ws_batch['A4'] = "Cookie Type:"
    ws_batch['A4'].font = Font(bold=True)
    cell_type = ws_batch['B4']
    cell_type.fill = input_fill
    cell_type.border = thin_border
    
    dv_batch = DataValidation(type="list", formula1=cookie_types_str, allow_blank=True)
    ws_batch.add_data_validation(dv_batch)
    dv_batch.add(cell_type)
    
    ws_batch['A5'] = "Quantity Needed:"
    ws_batch['A5'].font = Font(bold=True)
    ws_batch['B5'].fill = input_fill
    ws_batch['B5'].border = thin_border
    
    ws_batch['A6'] = "Cookie Size (g):"
    ws_batch['A6'].font = Font(bold=True)
    ws_batch['B6'].fill = input_fill
    ws_batch['B6'].border = thin_border
    
    # Calculated fields
    ws_batch['D4'] = "Base Yield:"
    ws_batch['D4'].font = Font(bold=True)
    ws_batch['E4'] = '=IFERROR(VLOOKUP(B4,\'Recipe Database\'!A:B,2,FALSE),"")'
    ws_batch['E4'].border = thin_border
    
    ws_batch['D5'] = "Base Size (g):"
    ws_batch['D5'].font = Font(bold=True)
    ws_batch['E5'] = '=IFERROR(VLOOKUP(B4,\'Recipe Database\'!A:C,3,FALSE),"")'
    ws_batch['E5'].border = thin_border
    
    ws_batch['D6'] = "Scale Factor:"
    ws_batch['D6'].font = Font(bold=True)
    ws_batch['E6'] = '=IF(OR(B5="",B6="",E4="",E5=""),"",(B5*B6)/(E4*E5))'
    ws_batch['E6'].border = thin_border
    ws_batch['E6'].fill = PatternFill(start_color="fff9c4", end_color="fff9c4", fill_type="solid")
    ws_batch['E6'].number_format = '0.00'
    ws_batch['F6'] = "x base recipe"
    ws_batch['F6'].font = Font(italic=True, color="666666")
    
    # Summary
    ws_batch['G4'] = "Total Dough:"
    ws_batch['G4'].font = Font(bold=True)
    ws_batch['H4'] = '=IF(OR(B5="",B6=""),"",B5*B6)'
    ws_batch['H4'].border = thin_border
    ws_batch['H4'].number_format = '#,##0 "g"'
    
    ws_batch['G5'] = "Total Cost:"
    ws_batch['G5'].font = Font(bold=True)
    ws_batch['H5'] = '=IF(H40="","",H40)'
    ws_batch['H5'].border = thin_border
    ws_batch['H5'].fill = money_fill
    ws_batch['H5'].number_format = '$#,##0.00'
    
    ws_batch['G6'] = "Cost/Cookie:"
    ws_batch['G6'].font = Font(bold=True)
    ws_batch['H6'] = '=IF(OR(H5="",B5=""),"",H5/B5)'
    ws_batch['H6'].border = thin_border
    ws_batch['H6'].fill = money_fill
    ws_batch['H6'].number_format = '$#,##0.00'
    
    # Ingredient breakdown headers
    batch_headers = ["Ingredient", "Base Amount", "Unit", "Scaled Amount", "Unit Price", "Ingredient Cost", "Category"]
    for col, header in enumerate(batch_headers, 1):
        cell = ws_batch.cell(row=8, column=col, value=header)
        cell.font = header_font
        cell.fill = subheader_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # List all possible ingredients with SUMIF formulas
    all_ingredients = [
        "Unsalted Butter", "Granulated Sugar", "Brown Sugar (Light)", "Brown Sugar (Dark)",
        "Eggs (large)", "Egg Yolk", "Vanilla Extract", "Sour Cream", "Cream Cheese",
        "All-Purpose Flour", "Cornstarch", "Dutch-Process Cocoa", "Old-Fashioned Oats",
        "Baking Powder", "Baking Soda", "Cream of Tartar",
        "Ground Cinnamon", "Salt (Kosher)", "Flaky Sea Salt", "Espresso Powder",
        "Chocolate Chips", "Dark Chocolate Chunks", "Raisins", "Walnuts (chopped)",
        "Lemon Juice (fresh)", "Lemon Zest"
    ]
    
    for i, ing in enumerate(all_ingredients):
        row = 9 + i
        
        # Ingredient name
        ws_batch.cell(row=row, column=1, value=ing).border = thin_border
        
        # Base amount - SUMIFS to get amount for selected cookie + this ingredient
        base_formula = f'=IFERROR(SUMIFS(\'Recipe Database\'!F:F,\'Recipe Database\'!A:A,$B$4,\'Recipe Database\'!E:E,A{row}),0)'
        cell_base = ws_batch.cell(row=row, column=2, value=base_formula)
        cell_base.border = thin_border
        cell_base.number_format = '#,##0.00'
        
        # Unit lookup
        unit_formula = f'=IFERROR(VLOOKUP(A{row},\'Ingredient Prices\'!A:B,2,FALSE),"")'
        ws_batch.cell(row=row, column=3, value=unit_formula).border = thin_border
        
        # Scaled amount = base * scale factor
        scaled_formula = f'=IF(OR($E$6="",B{row}=0),"",ROUND(B{row}*$E$6,1))'
        cell_scaled = ws_batch.cell(row=row, column=4, value=scaled_formula)
        cell_scaled.border = thin_border
        cell_scaled.fill = PatternFill(start_color="fff3e0", end_color="fff3e0", fill_type="solid")
        cell_scaled.number_format = '#,##0.0'
        
        # Unit price lookup
        price_formula = f'=IFERROR(VLOOKUP(A{row},\'Ingredient Prices\'!A:E,5,FALSE),0)'
        cell_price = ws_batch.cell(row=row, column=5, value=price_formula)
        cell_price.border = thin_border
        cell_price.number_format = '$#,##0.0000'
        
        # Ingredient cost = scaled amount * unit price
        cost_formula = f'=IF(D{row}="",0,D{row}*E{row})'
        cell_cost = ws_batch.cell(row=row, column=6, value=cost_formula)
        cell_cost.border = thin_border
        cell_cost.fill = money_fill
        cell_cost.number_format = '$#,##0.00'
        
        # Category lookup
        cat_formula = f'=IFERROR(INDEX(\'Recipe Database\'!H:H,MATCH(1,(\'Recipe Database\'!A:A=$B$4)*(\'Recipe Database\'!E:E=A{row}),0)),"")'
        ws_batch.cell(row=row, column=7, value=cat_formula).border = thin_border
    
    # Total row
    total_row = 9 + len(all_ingredients) + 1
    ws_batch.cell(row=total_row, column=5, value="TOTAL COST:").font = Font(bold=True, size=12)
    ws_batch.cell(row=total_row, column=6, value=f"=SUM(F9:F{total_row-1})").font = Font(bold=True, size=12)
    ws_batch.cell(row=total_row, column=6).fill = PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid")
    ws_batch.cell(row=total_row, column=6).number_format = '$#,##0.00'
    ws_batch.cell(row=total_row, column=6).border = thin_border
    
    # Update H5 to reference correct total row
    ws_batch['H5'] = f'=IF(F{total_row}="","",F{total_row})'
    
    # Column widths
    ws_batch.column_dimensions['A'].width = 25
    ws_batch.column_dimensions['B'].width = 14
    ws_batch.column_dimensions['C'].width = 8
    ws_batch.column_dimensions['D'].width = 16
    ws_batch.column_dimensions['E'].width = 14
    ws_batch.column_dimensions['F'].width = 16
    ws_batch.column_dimensions['G'].width = 12
    ws_batch.column_dimensions['H'].width = 14
    
    # =========================================================================
    # SHEET 4: ORDER CALCULATOR
    # =========================================================================
    ws_order = wb.create_sheet("Order Calculator")
    
    # Header
    ws_order.merge_cells('A1:I1')
    ws_order['A1'] = "🍪 MIDNIGHT DOUGH - ORDER CALCULATOR"
    ws_order['A1'].font = Font(color="FFFFFF", bold=True, size=16)
    ws_order['A1'].fill = header_fill
    ws_order['A1'].alignment = Alignment(horizontal='center')
    
    # Instructions
    ws_order.merge_cells('A2:I2')
    ws_order['A2'] = "Add multiple orders. Each row calculates batches needed and estimated ingredient cost."
    ws_order['A2'].font = Font(italic=True, size=10)
    ws_order['A2'].alignment = Alignment(horizontal='center')
    
    # Order entry headers
    order_headers = ["#", "Cookie Type", "Qty", "Size (g)", "Total Dough (g)", 
                     "Scale Factor", "Est. Cost ($)", "Cost/Cookie", "Notes"]
    for col, header in enumerate(order_headers, 1):
        cell = ws_order.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Create dropdown for cookie types
    dv = DataValidation(type="list", formula1=cookie_types_str, allow_blank=True)
    dv.error = "Please select a valid cookie type"
    dv.errorTitle = "Invalid Cookie Type"
    ws_order.add_data_validation(dv)
    
    # Add 20 order rows with formulas
    for i in range(20):
        row = 5 + i
        order_num = i + 1
        
        # Order number
        ws_order.cell(row=row, column=1, value=order_num).border = thin_border
        
        # Cookie type (dropdown)
        cell_cookie = ws_order.cell(row=row, column=2)
        cell_cookie.border = thin_border
        cell_cookie.fill = input_fill
        dv.add(cell_cookie)
        
        # Quantity (user input)
        cell_qty = ws_order.cell(row=row, column=3)
        cell_qty.border = thin_border
        cell_qty.fill = input_fill
        
        # Cookie size (user input)
        cell_size = ws_order.cell(row=row, column=4)
        cell_size.border = thin_border
        cell_size.fill = input_fill
        
        # Total dough needed formula
        cell_dough = ws_order.cell(row=row, column=5, value=f'=IF(C{row}="","",C{row}*D{row})')
        cell_dough.border = thin_border
        cell_dough.number_format = '#,##0'
        
        # Scale factor
        scale_formula = f'=IF(B{row}="","",E{row}/VLOOKUP(B{row},\'Recipe Database\'!A:D,4,FALSE))'
        cell_scale = ws_order.cell(row=row, column=6, value=scale_formula)
        cell_scale.border = thin_border
        cell_scale.number_format = '0.00'
        
        # Est cost formula - calculate from recipe cost lookup
        # First we need to know the base cost per batch, then multiply by scale factor
        cost_formula = f'=IFERROR(IF(F{row}="","",F{row}*SUMPRODUCT((\'Recipe Database\'!A$4:A$100=B{row})*(\'Recipe Database\'!F$4:F$100)*VLOOKUP(\'Recipe Database\'!E$4:E$100,\'Ingredient Prices\'!A:E,5,FALSE))),"")'
        cell_cost = ws_order.cell(row=row, column=7, value=cost_formula)
        cell_cost.border = thin_border
        cell_cost.number_format = '$#,##0.00'
        cell_cost.fill = money_fill
        
        # Price per cookie
        price_per = f'=IF(OR(G{row}="",C{row}=""),"",G{row}/C{row})'
        cell_ppc = ws_order.cell(row=row, column=8, value=price_per)
        cell_ppc.border = thin_border
        cell_ppc.number_format = '$#,##0.00'
        cell_ppc.fill = money_fill
        
        # Notes
        ws_order.cell(row=row, column=9).border = thin_border
    
    # Totals row
    total_row = 26
    ws_order.cell(row=total_row, column=1, value="TOTALS").font = Font(bold=True)
    ws_order.cell(row=total_row, column=3, value="=SUM(C5:C25)").font = Font(bold=True)
    ws_order.cell(row=total_row, column=5, value="=SUM(E5:E25)").font = Font(bold=True)
    ws_order.cell(row=total_row, column=5).number_format = '#,##0'
    ws_order.cell(row=total_row, column=7, value="=SUM(G5:G25)").font = Font(bold=True)
    ws_order.cell(row=total_row, column=7).number_format = '$#,##0.00'
    ws_order.cell(row=total_row, column=7).fill = PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid")
    
    for col in range(1, 10):
        ws_order.cell(row=total_row, column=col).border = thin_border
    
    # Set column widths
    ws_order.column_dimensions['A'].width = 5
    ws_order.column_dimensions['B'].width = 22
    ws_order.column_dimensions['C'].width = 8
    ws_order.column_dimensions['D'].width = 10
    ws_order.column_dimensions['E'].width = 16
    ws_order.column_dimensions['F'].width = 12
    ws_order.column_dimensions['G'].width = 14
    ws_order.column_dimensions['H'].width = 12
    ws_order.column_dimensions['I'].width = 20
    
    # =========================================================================
    # SHEET 5: SHOPPING LIST
    # =========================================================================
    ws_shopping = wb.create_sheet("Shopping List")
    
    # Header
    ws_shopping.merge_cells('A1:H1')
    ws_shopping['A1'] = "🛒 INGREDIENT SHOPPING LIST"
    ws_shopping['A1'].font = Font(color="FFFFFF", bold=True, size=16)
    ws_shopping['A1'].fill = header_fill
    ws_shopping['A1'].alignment = Alignment(horizontal='center')
    
    ws_shopping.merge_cells('A2:H2')
    ws_shopping['A2'] = "Enter amounts needed → see packages to buy and total cost"
    ws_shopping['A2'].font = Font(italic=True, size=10)
    ws_shopping['A2'].alignment = Alignment(horizontal='center')
    
    shop_headers = ["Ingredient", "Amount Needed", "Unit", "Pkg Size", "Pkgs to Buy", "Pkg Price", "Total Cost", "Notes"]
    for col, header in enumerate(shop_headers, 1):
        cell = ws_shopping.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    for i, ing in enumerate(all_ingredients):
        row = 5 + i
        ws_shopping.cell(row=row, column=1, value=ing).border = thin_border
        
        # Amount needed (user enters)
        cell_needed = ws_shopping.cell(row=row, column=2)
        cell_needed.border = thin_border
        cell_needed.fill = input_fill
        
        # Unit lookup
        unit_formula = f"=IFERROR(VLOOKUP(A{row},'Ingredient Prices'!A:B,2,FALSE),\"\")"
        ws_shopping.cell(row=row, column=3, value=unit_formula).border = thin_border
        
        # Package size lookup
        size_formula = f"=IFERROR(VLOOKUP(A{row},'Ingredient Prices'!A:C,3,FALSE),\"\")"
        ws_shopping.cell(row=row, column=4, value=size_formula).border = thin_border
        
        # Packages to buy
        pkg_formula = f'=IF(B{row}="","",ROUNDUP(B{row}/D{row},0))'
        ws_shopping.cell(row=row, column=5, value=pkg_formula).border = thin_border
        
        # Package price lookup
        price_formula = f"=IFERROR(VLOOKUP(A{row},'Ingredient Prices'!A:D,4,FALSE),\"\")"
        cell_price = ws_shopping.cell(row=row, column=6, value=price_formula)
        cell_price.border = thin_border
        cell_price.number_format = '$#,##0.00'
        
        # Total cost
        total_formula = f'=IF(E{row}="","",E{row}*F{row})'
        cell_total = ws_shopping.cell(row=row, column=7, value=total_formula)
        cell_total.border = thin_border
        cell_total.number_format = '$#,##0.00'
        cell_total.fill = money_fill
        
        # Notes
        ws_shopping.cell(row=row, column=8).border = thin_border
    
    # Grand total
    total_row = 5 + len(all_ingredients)
    ws_shopping.cell(row=total_row, column=6, value="GRAND TOTAL:").font = Font(bold=True)
    ws_shopping.cell(row=total_row, column=7, value=f"=SUM(G5:G{total_row-1})").font = Font(bold=True)
    ws_shopping.cell(row=total_row, column=7).number_format = '$#,##0.00'
    ws_shopping.cell(row=total_row, column=7).fill = PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid")
    
    # Column widths
    ws_shopping.column_dimensions['A'].width = 25
    ws_shopping.column_dimensions['B'].width = 14
    ws_shopping.column_dimensions['C'].width = 8
    ws_shopping.column_dimensions['D'].width = 10
    ws_shopping.column_dimensions['E'].width = 12
    ws_shopping.column_dimensions['F'].width = 12
    ws_shopping.column_dimensions['G'].width = 12
    ws_shopping.column_dimensions['H'].width = 20
    
    # =========================================================================
    # SHEET 6: QUICK REFERENCE
    # =========================================================================
    ws_ref = wb.create_sheet("Quick Reference")
    
    ws_ref.merge_cells('A1:D1')
    ws_ref['A1'] = "📋 QUICK REFERENCE - COOKIE SIZES & BAKE TIMES"
    ws_ref['A1'].font = Font(color="FFFFFF", bold=True, size=16)
    ws_ref['A1'].fill = header_fill
    ws_ref['A1'].alignment = Alignment(horizontal='center')
    
    ref_headers = ["Cookie", "Recommended Size (g)", "Oven Temp", "Bake Time"]
    for col, header in enumerate(ref_headers, 1):
        cell = ws_ref.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = subheader_fill
        cell.border = thin_border
    
    ref_data = [
        ("Sugar Cookie", "45-50g", "350°F (175°C)", "9-11 min"),
        ("Snickerdoodle", "40g (sm) / 70-80g (lg)", "375°F (190°C)", "10 min"),
        ("Dark Chocolate Chip", "110-120g", "375°F (190°C)", "12-14 min"),
        ("Chocolate Chip", "40-70g / 80-100g (lg)", "350°F (175°C)", "10-12 min"),
        ("Lemon Sugar", "30g / 45-55g (bakery)", "350°F (175°C)", "10-12 min"),
        ("Oatmeal Raisin", "45-55g", "350°F (175°C)", "11-13 min"),
    ]
    
    for i, data in enumerate(ref_data, 4):
        for col, val in enumerate(data, 1):
            cell = ws_ref.cell(row=i, column=col, value=val)
            cell.border = thin_border
    
    ws_ref.column_dimensions['A'].width = 22
    ws_ref.column_dimensions['B'].width = 25
    ws_ref.column_dimensions['C'].width = 18
    ws_ref.column_dimensions['D'].width = 12
    
    # Save the workbook
    filename = "midnight_dough_cookie_calculator.xlsx"
    wb.save(filename)
    print(f"✅ Calculator created successfully: {filename}")
    print("\n📊 SHEETS:")
    print("  1. Ingredient Prices - Edit package prices (per-unit auto-calculates)")
    print("  2. Recipe Database   - All 6 cookie recipes")
    print("  3. Batch Calculator  - Select cookie, qty, size → full breakdown")
    print("  4. Order Calculator  - Multiple orders with totals")
    print("  5. Shopping List     - Calculate packages to buy")
    print("  6. Quick Reference   - Sizes and bake times")
    
    return filename

if __name__ == "__main__":
    create_cookie_calculator()
